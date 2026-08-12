from datetime import datetime
from io import BytesIO

from openpyxl.cell.cell import MergedCell
from openpyxl import load_workbook

from app.schemas.measurement import MeasurementResponse
from app.schemas.recipe import RecipeBase, SectionParams
from app.services.measurement_export import (
    build_measurements_xlsx,
    format_recorded_at,
    sanitize_export_filename,
    sanitize_sheet_title,
    _limits_for_metric,
    _tier_totals_from_batch_totals,
)

_EXPORT_TEST_RECIPE = RecipeBase(
    name="测试配方",
    batch_size=4,
    temperature={"min": 20, "max": 30},
    weight={"min": 100, "max": 150},
    length={"min": 90, "max": 110},
    width={"min": 40, "max": 55},
    height={"min": 25, "max": 35},
    water_cut_width={"min": 40, "max": 50},
    enable_water_cut=True,
    enable_bottom_measurement=True,
    bottom_params=SectionParams(
        batch_size=2,
        temperature={"min": 18, "max": 24},
        weight={"min": 80, "max": 95},
        length={"min": 85, "max": 100},
        width={"min": 35, "max": 45},
        height={"min": 20, "max": 28},
        water_cut_width={"min": 30, "max": 40},
    ),
    enable_middle_measurement=True,
    middle_params=SectionParams(
        batch_size=2,
        temperature={"min": 19, "max": 25},
        weight={"min": 90, "max": 105},
        length={"min": 88, "max": 102},
        width={"min": 38, "max": 48},
        height={"min": 22, "max": 30},
        water_cut_width={"min": 32, "max": 42},
    ),
)


def _row_values(sheet, row: int, last_col: int) -> list:
    return [sheet.cell(row=row, column=col).value for col in range(1, last_col + 1)]


def test_sanitize_export_filename_replaces_illegal_chars() -> None:
    assert sanitize_export_filename('数据汇总:2026/7/11') == "数据汇总_2026_7_11.xlsx"


def test_sanitize_export_filename_converts_csv_suffix() -> None:
    assert sanitize_export_filename("数据汇总.csv") == "数据汇总.xlsx"


def test_sanitize_sheet_title() -> None:
    assert sanitize_sheet_title("成品-温度") == "成品-温度"
    assert len(sanitize_sheet_title("x" * 40)) == 31


def test_build_measurements_xlsx_creates_metric_sheets_per_type() -> None:
    records = [
        MeasurementResponse(
            id="1",
            batch_id=1,
            recipe_id="standardC",
            record_type="product",
            slot_index=0,
            sample_name="样品-成品-1",
            temperature="24.5",
            weight="128.3",
            length="101.2",
            width="49.5",
            height="29.8",
            water_cut_width="0",
            preview_name=None,
            recorded_at=datetime(2026, 6, 22, 16, 32, 56),
        ),
        MeasurementResponse(
            id="2",
            batch_id=1,
            recipe_id="standardC",
            record_type="bottom",
            slot_index=0,
            sample_name="样品-底片-1",
            temperature="22.1",
            weight="88.3",
            length="91.2",
            width="45.5",
            height="28.8",
            water_cut_width="0",
            preview_name=None,
            recorded_at=datetime(2026, 6, 22, 16, 33, 10),
        ),
    ]
    content = build_measurements_xlsx(records, enable_water_cut=False)
    workbook = load_workbook(BytesIO(content))

    assert workbook.sheetnames[0] == "底稿"
    draft = workbook["底稿"]
    assert _row_values(draft, 1, 8) == [
        "批次号",
        "名称",
        "温度(°C)",
        "重量(g)",
        "长(mm)",
        "宽(mm)",
        "高(mm)",
        "时间",
    ]
    assert draft.max_row == 3
    assert _row_values(draft, 2, 8) == [
        1,
        "样品-底片-1",
        "22.1",
        "88.3",
        "91.2",
        "45.5",
        "28.8",
        "2026/06/22 16:33:10",
    ]
    assert _row_values(draft, 3, 8) == [
        1,
        "样品-成品-1",
        "24.5",
        "128.3",
        "101.2",
        "49.5",
        "29.8",
        "2026/06/22 16:32:56",
    ]

    assert "成品-温度" in workbook.sheetnames
    assert "成品-重量" in workbook.sheetnames
    assert "底片-温度" in workbook.sheetnames
    assert "底片-重量" in workbook.sheetnames
    assert "成品-水切宽度" not in workbook.sheetnames

    product_temperature = workbook["成品-温度"]
    assert _row_values(product_temperature, 1, 5) == [
        "批次号",
        "开始时间",
        "温度1",
        "单批温度",
        "单打温度",
    ]
    assert _row_values(product_temperature, 2, 5) == [
        1,
        "2026/06/22 16:32:56",
        24.5,
        24.5,
        None,
    ]


def test_build_measurements_xlsx_groups_batch_into_one_row_with_sum() -> None:
    records = [
        MeasurementResponse(
            id="1",
            batch_id=1,
            recipe_id="standardC",
            record_type="product",
            slot_index=0,
            sample_name="样品-成品-1",
            temperature="24.5",
            weight="100",
            length="0",
            width="0",
            height="0",
            water_cut_width="0",
            preview_name=None,
            recorded_at=datetime(2026, 6, 22, 16, 32, 56),
        ),
        MeasurementResponse(
            id="2",
            batch_id=1,
            recipe_id="standardC",
            record_type="product",
            slot_index=1,
            sample_name="样品-成品-2",
            temperature="25.5",
            weight="110",
            length="0",
            width="0",
            height="0",
            water_cut_width="0",
            preview_name=None,
            recorded_at=datetime(2026, 6, 22, 16, 33, 10),
        ),
        MeasurementResponse(
            id="3",
            batch_id=2,
            recipe_id="standardC",
            record_type="product",
            slot_index=0,
            sample_name="样品-成品-3",
            temperature="20",
            weight="90",
            length="0",
            width="0",
            height="0",
            water_cut_width="0",
            preview_name=None,
            recorded_at=datetime(2026, 6, 22, 17, 0, 0),
        ),
    ]
    workbook = load_workbook(BytesIO(build_measurements_xlsx(records, enable_water_cut=False)))
    temperature_sheet = workbook["成品-温度"]
    weight_sheet = workbook["成品-重量"]

    assert _row_values(temperature_sheet, 1, 6) == [
        "批次号",
        "开始时间",
        "温度1",
        "温度2",
        "单批温度",
        "单打温度",
    ]
    assert _row_values(temperature_sheet, 2, 6) == [
        1,
        "2026/06/22 16:32:56",
        24.5,
        25.5,
        50,
        None,
    ]
    assert _row_values(temperature_sheet, 3, 6) == [
        2,
        "2026/06/22 17:00:00",
        20,
        None,
        20,
        70,
    ]

    assert _row_values(weight_sheet, 1, 6) == [
        "批次号",
        "开始时间",
        "重量1",
        "重量2",
        "单批重量",
        "单打重量",
    ]
    assert _row_values(weight_sheet, 2, 6) == [
        1,
        "2026/06/22 16:32:56",
        100,
        110,
        210,
        None,
    ]
    assert _row_values(weight_sheet, 3, 6) == [
        2,
        "2026/06/22 17:00:00",
        90,
        None,
        90,
        300,
    ]


def test_build_measurements_xlsx_includes_water_cut_for_product_when_enabled() -> None:
    records = [
        MeasurementResponse(
            id="1",
            batch_id=1,
            recipe_id="standardC",
            record_type="product",
            slot_index=0,
            sample_name="样品-成品-1",
            temperature="24.5",
            weight="128.3",
            length="101.2",
            width="49.5",
            height="29.8",
            water_cut_width="42.5",
            preview_name=None,
            recorded_at=datetime(2026, 6, 22, 16, 32, 56),
        )
    ]
    workbook = load_workbook(
        BytesIO(build_measurements_xlsx(records, enable_water_cut=True))
    )
    assert "成品-水切宽度" in workbook.sheetnames
    assert workbook.sheetnames[0] == "底稿"
    draft = workbook["底稿"]
    assert _row_values(draft, 1, 9) == [
        "批次号",
        "名称",
        "温度(°C)",
        "重量(g)",
        "长(mm)",
        "宽(mm)",
        "高(mm)",
        "水切宽度(mm)",
        "时间",
    ]
    assert _row_values(draft, 2, 9) == [
        1,
        "样品-成品-1",
        "24.5",
        "128.3",
        "101.2",
        "49.5",
        "29.8",
        "42.5",
        "2026/06/22 16:32:56",
    ]
    assert _row_values(workbook["成品-水切宽度"], 1, 5) == [
        "批次号",
        "开始时间",
        "水切宽度1",
        "单批水切宽度",
        "单打水切宽度",
    ]
    assert _row_values(workbook["成品-水切宽度"], 2, 5) == [
        1,
        "2026/06/22 16:32:56",
        42.5,
        42.5,
        None,
    ]


def test_format_recorded_at() -> None:
    assert format_recorded_at(datetime(2026, 6, 22, 16, 32, 56)) == "2026/06/22 16:32:56"


def test_build_measurements_xlsx_appends_stats_table_with_formulas() -> None:
    records = [
        MeasurementResponse(
            id="1",
            batch_id=1,
            recipe_id="standardC",
            record_type="product",
            slot_index=0,
            sample_name="样品-成品-1",
            temperature="24.5",
            weight="100",
            length="0",
            width="0",
            height="0",
            water_cut_width="0",
            preview_name=None,
            recorded_at=datetime(2026, 6, 22, 16, 32, 56),
        ),
        MeasurementResponse(
            id="2",
            batch_id=2,
            recipe_id="standardC",
            record_type="product",
            slot_index=0,
            sample_name="样品-成品-2",
            temperature="26.5",
            weight="110",
            length="0",
            width="0",
            height="0",
            water_cut_width="0",
            preview_name=None,
            recorded_at=datetime(2026, 6, 22, 17, 0, 0),
        ),
    ]
    workbook = load_workbook(
        BytesIO(
            build_measurements_xlsx(
                records,
                enable_water_cut=False,
                recipe=_EXPORT_TEST_RECIPE,
            )
        )
    )
    sheet = workbook["成品-重量"]

    title_row = 5  # header + 2 data rows + blank row
    stats_start = title_row + 1
    tier_value_col = 9
    tier_input_col = 8
    batch_value_col = 13
    sample_value_col = 17
    assert sheet.cell(row=title_row, column=7).value == "测试配方-单打重量"
    assert sheet.cell(row=title_row, column=11).value == "测试配方-单批重量"
    assert sheet.cell(row=title_row, column=15).value == "测试配方-单值重量"
    assert sheet.cell(row=stats_start, column=7).value == "公差上限 USL"
    assert sheet.cell(row=stats_start, column=tier_input_col).value == 1800
    assert sheet.cell(row=stats_start + 1, column=tier_input_col).value == 1200
    batch_input_col = 12
    sample_input_col = 16
    assert sheet.cell(row=stats_start, column=batch_input_col).value == 900
    assert sheet.cell(row=stats_start + 1, column=batch_input_col).value == 600
    assert sheet.cell(row=stats_start, column=sample_input_col).value == 150
    assert sheet.cell(row=stats_start + 1, column=sample_input_col).value == 100
    assert sheet.cell(row=stats_start, column=tier_value_col).value == "=MAX(E2:E3)"
    assert sheet.cell(row=stats_start + 1, column=tier_value_col).value == "=MIN(E2:E3)"
    assert sheet.cell(row=stats_start, column=11).value == "公差上限 USL"
    assert sheet.cell(row=stats_start, column=batch_value_col).value == "=MAX(D2:D3)"
    assert sheet.cell(row=stats_start, column=15).value == "公差上限 USL"
    assert sheet.cell(row=stats_start, column=sample_value_col).value == "=MAX(C2:C3)"
    assert sheet.cell(row=stats_start + 4, column=tier_value_col).value == "=AVERAGE(E2:E3)"
    assert sheet.cell(row=stats_start + 5, column=tier_value_col).value == "=STDEV(E2:E3)"
    assert sheet.cell(row=stats_start + 2, column=tier_value_col).value == "=(H6+H7)/2"
    assert sheet.cell(row=stats_start + 3, column=tier_value_col).value == "=H6-H7"
    assert sheet.cell(row=stats_start + 8, column=tier_value_col).value == "=(H6-I10)/(3*I11)"
    assert sheet.cell(row=stats_start + 10, column=7).value == "CPK"
    assert sheet.cell(row=stats_start + 10, column=tier_value_col).value == "=MIN(I14,I15)"
    assert sheet.cell(row=stats_start + 14, column=tier_value_col).value == "=I18*(1-ABS(I19))"


def test_tier_totals_from_batch_totals_pairs_and_discards_trailing_odd_batch() -> None:
    assert _tier_totals_from_batch_totals([50, 20, 30]) == [None, 70, None]
    assert _tier_totals_from_batch_totals([10, 20, 30, 40]) == [None, 30, None, 70]
    assert _tier_totals_from_batch_totals([10]) == [None]


def test_build_measurements_xlsx_stats_table_has_black_border() -> None:
    records = [
        MeasurementResponse(
            id="1",
            batch_id=1,
            recipe_id="standardC",
            record_type="product",
            slot_index=0,
            sample_name="样品-成品-1",
            temperature="24.5",
            weight="100",
            length="0",
            width="0",
            height="0",
            water_cut_width="0",
            preview_name=None,
            recorded_at=datetime(2026, 6, 22, 16, 32, 56),
        ),
        MeasurementResponse(
            id="2",
            batch_id=2,
            recipe_id="standardC",
            record_type="product",
            slot_index=0,
            sample_name="样品-成品-2",
            temperature="26.5",
            weight="110",
            length="0",
            width="0",
            height="0",
            water_cut_width="0",
            preview_name=None,
            recorded_at=datetime(2026, 6, 22, 17, 0, 0),
        ),
    ]
    workbook = load_workbook(
        BytesIO(
            build_measurements_xlsx(
                records,
                enable_water_cut=False,
                recipe=_EXPORT_TEST_RECIPE,
            )
        )
    )
    sheet = workbook["成品-重量"]
    title_row = 5
    stats_start = title_row + 1
    table_ranges = (
        range(7, 10),
        range(11, 14),
        range(15, 18),
    )
    stats_end = stats_start + 14
    for cols in table_ranges:
        for row in range(title_row, stats_end + 1):
            for col in cols:
                cell = sheet.cell(row=row, column=col)
                if isinstance(cell, MergedCell):
                    continue
                border = cell.border
                assert border.left.style == "thin"
                assert border.right.style == "thin"
                assert border.top.style == "thin"
                assert border.bottom.style == "thin"


def test_limits_for_metric_uses_section_params() -> None:
    assert _limits_for_metric(_EXPORT_TEST_RECIPE, "product", "weight") == (100, 150)
    assert _limits_for_metric(_EXPORT_TEST_RECIPE, "bottom", "weight") == (80, 95)
    assert _limits_for_metric(_EXPORT_TEST_RECIPE, "middle", "temperature") == (19, 25)


def test_build_measurements_xlsx_stats_tables_use_data_derived_limits() -> None:
    records = [
        MeasurementResponse(
            id="1",
            batch_id=1,
            recipe_id="standardC",
            record_type="bottom",
            slot_index=0,
            sample_name="样品-底片-1",
            temperature="22.1",
            weight="88.3",
            length="91.2",
            width="45.5",
            height="28.8",
            water_cut_width="0",
            preview_name=None,
            recorded_at=datetime(2026, 6, 22, 16, 33, 10),
        ),
    ]
    workbook = load_workbook(
        BytesIO(
            build_measurements_xlsx(
                records,
                enable_water_cut=False,
                recipe=_EXPORT_TEST_RECIPE,
            )
        )
    )
    sheet = workbook["底片-重量"]
    stats_start = 5  # header + 1 data row + blank row + title row
    assert sheet.cell(row=stats_start - 1, column=7).value == "测试配方-单打重量"
    assert sheet.cell(row=stats_start, column=8).value == 1140
    assert sheet.cell(row=stats_start + 1, column=8).value == 960
    assert sheet.cell(row=stats_start, column=12).value == 570
    assert sheet.cell(row=stats_start + 1, column=12).value == 480
    assert sheet.cell(row=stats_start, column=16).value == 95
    assert sheet.cell(row=stats_start + 1, column=16).value == 80
    assert (
        sheet.cell(row=stats_start, column=13).value
        == "=MAX(D2:D2)"
    )


def test_build_measurements_xlsx_round_bread_exports_diameter_not_length_width() -> None:
    round_recipe = _EXPORT_TEST_RECIPE.model_copy(update={"enable_round_bread": True})
    records = [
        MeasurementResponse(
            id="1",
            batch_id=1,
            recipe_id="roundA",
            record_type="product",
            slot_index=0,
            sample_name="圆包-成品-1",
            temperature="24.5",
            weight="128.3",
            length="100.0",
            width="100.0",
            height="29.8",
            water_cut_width="0",
            preview_name=None,
            recorded_at=datetime(2026, 6, 22, 16, 32, 56),
        ),
        MeasurementResponse(
            id="2",
            batch_id=1,
            recipe_id="roundA",
            record_type="bottom",
            slot_index=0,
            sample_name="圆包-底片-1",
            temperature="22.1",
            weight="88.3",
            length="91.2",
            width="91.2",
            height="28.8",
            water_cut_width="0",
            preview_name=None,
            recorded_at=datetime(2026, 6, 22, 16, 33, 10),
        ),
        MeasurementResponse(
            id="3",
            batch_id=1,
            recipe_id="roundA",
            record_type="middle",
            slot_index=0,
            sample_name="圆包-中片-1",
            temperature="23.0",
            weight="90.0",
            length="95.0",
            width="95.0",
            height="27.0",
            water_cut_width="0",
            preview_name=None,
            recorded_at=datetime(2026, 6, 22, 16, 34, 0),
        ),
    ]
    workbook = load_workbook(
        BytesIO(
            build_measurements_xlsx(
                records,
                enable_water_cut=False,
                recipe=round_recipe,
            )
        )
    )

    draft = workbook["底稿"]
    assert _row_values(draft, 1, 7) == [
        "批次号",
        "名称",
        "温度(°C)",
        "重量(g)",
        "直径(mm)",
        "高(mm)",
        "时间",
    ]
    assert _row_values(draft, 2, 7) == [
        1,
        "圆包-中片-1",
        "23.0",
        "90.0",
        "95.0",
        "27.0",
        "2026/06/22 16:34:00",
    ]

    assert "成品-直径" in workbook.sheetnames
    assert "底片-直径" in workbook.sheetnames
    assert "中片-直径" in workbook.sheetnames
    assert "成品-长" not in workbook.sheetnames
    assert "成品-宽" not in workbook.sheetnames
    assert "底片-长" not in workbook.sheetnames
    assert "底片-宽" not in workbook.sheetnames

    product_diameter = workbook["成品-直径"]
    assert _row_values(product_diameter, 1, 5) == [
        "批次号",
        "开始时间",
        "直径1",
        "单批直径",
        "单打直径",
    ]
    assert _row_values(product_diameter, 2, 5) == [
        1,
        "2026/06/22 16:32:56",
        100.0,
        100.0,
        None,
    ]

    bottom_diameter = workbook["底片-直径"]
    stats_start = 5
    assert bottom_diameter.cell(row=stats_start - 1, column=7).value == "测试配方-单打直径"
    assert bottom_diameter.cell(row=stats_start, column=8).value == 1200
    assert bottom_diameter.cell(row=stats_start + 1, column=8).value == 1020
