from __future__ import annotations

import io
import re
from collections.abc import Callable
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.schemas.measurement import MeasurementResponse, RecordType
from app.schemas.recipe import RecipeBase, SectionParams

DRAFT_SHEET_TITLE = "底稿"

RECORD_TYPE_LABELS: dict[RecordType, str] = {
    "product": "成品",
    "bottom": "底片",
    "middle": "中片",
}

METRIC_FIELDS: tuple[tuple[str, str, str], ...] = (
    ("temperature", "温度", "温度(°C)"),
    ("weight", "重量", "重量(g)"),
    ("length", "长", "长(mm)"),
    ("width", "宽", "宽(mm)"),
    ("height", "高", "高(mm)"),
    ("water_cut_width", "水切宽度", "水切宽度(mm)"),
)

_WINDOWS_ILLEGAL_FILENAME = re.compile(r'[<>:"/\\|?*]')
_SHEET_ILLEGAL_CHARS = re.compile(r"[\\/*?:\[\]]")


def sanitize_export_filename(filename: str) -> str:
    cleaned = _WINDOWS_ILLEGAL_FILENAME.sub("_", filename.strip())
    if cleaned.lower().endswith(".csv"):
        cleaned = cleaned[:-4]
    if not cleaned.lower().endswith(".xlsx"):
        cleaned = f"{cleaned}.xlsx"
    return cleaned


def sanitize_sheet_title(title: str) -> str:
    cleaned = _SHEET_ILLEGAL_CHARS.sub("_", title.strip())
    return cleaned[:31] or "Sheet"


def format_recorded_at(dt: datetime) -> str:
    """近似前端 toLocaleString('zh-CN') 的展示格式。"""
    return dt.strftime("%Y/%m/%d %H:%M:%S")


def _record_sort_key(record: MeasurementResponse) -> tuple[int, int, datetime]:
    return (
        record.batch_id or 0,
        record.slot_index,
        record.recorded_at,
    )


def _metric_value(record: MeasurementResponse, field: str) -> str:
    return getattr(record, field)


def _parse_numeric(value: str) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _format_sum(total: float) -> str | int | float:
    if total.is_integer():
        return int(total)
    return round(total, 4)


def _tier_totals_from_batch_totals(
    batch_totals: list[float],
) -> list[float | None]:
    """Pair consecutive batch totals; write tier sum on the second row of each pair."""
    tier_values: list[float | None] = [None] * len(batch_totals)
    for index in range(0, len(batch_totals) - 1, 2):
        tier_values[index + 1] = batch_totals[index] + batch_totals[index + 1]
    return tier_values


def _group_records_by_batch(
    records: list[MeasurementResponse],
) -> list[tuple[int, list[MeasurementResponse]]]:
    batches: dict[int, list[MeasurementResponse]] = {}
    for record in records:
        batch_key = record.batch_id or 0
        batches.setdefault(batch_key, []).append(record)

    grouped: list[tuple[int, list[MeasurementResponse]]] = []
    for batch_id in sorted(batches):
        batch_records = sorted(
            batches[batch_id],
            key=lambda record: (record.slot_index, record.recorded_at),
        )
        grouped.append((batch_id, batch_records))
    return grouped


def _export_metric_fields(
    *,
    enable_round_bread: bool,
) -> tuple[tuple[str, str, str], ...]:
    """Return metric columns/sheets to export, relabeling length as diameter for round bread."""
    fields: list[tuple[str, str, str]] = []
    for field, short_label, header in METRIC_FIELDS:
        if enable_round_bread and field == "width":
            continue
        if enable_round_bread and field == "length":
            fields.append((field, "直径", "直径(mm)"))
        else:
            fields.append((field, short_label, header))
    return tuple(fields)


def _should_include_metric(
    record_type: RecordType,
    field: str,
    *,
    enable_water_cut: bool,
    enable_round_bread: bool,
) -> bool:
    if enable_round_bread and field == "width":
        return False
    if field != "water_cut_width":
        return True
    return enable_water_cut and record_type == "product"


def _batch_start_time(batch_records: list[MeasurementResponse]) -> str:
    earliest = min(record.recorded_at for record in batch_records)
    return format_recorded_at(earliest)


def _cell_ref(row: int, column: int) -> str:
    return f"{get_column_letter(column)}{row}"


def _section_params_for_type(
    recipe: RecipeBase,
    record_type: RecordType,
) -> RecipeBase | SectionParams:
    if record_type == "bottom":
        return recipe.bottom_params
    if record_type == "middle":
        return recipe.middle_params
    return recipe


def _limits_for_metric(
    recipe: RecipeBase,
    record_type: RecordType,
    field: str,
) -> tuple[float, float]:
    """Return (LSL/min, USL/max) for the recipe section and metric field."""
    section = _section_params_for_type(recipe, record_type)
    spec = getattr(section, field)
    return spec.min, spec.max


def _metric_cell_value(raw: str) -> int | float | None:
    """Write parseable measurements as numbers so Excel stats formulas work."""
    if not str(raw).strip():
        return None
    return _format_sum(_parse_numeric(raw))


def _mean_formula(data_range: str) -> str:
    return f"=AVERAGE({data_range})"


def _stdev_formula(data_range: str) -> str:
    """STDEV (sample σ) is supported by Excel, WPS, and Apple Numbers; blank cells are ignored."""
    return f"=STDEV({data_range})"


def _max_formula(data_range: str) -> str:
    return f"=MAX({data_range})"


def _min_formula(data_range: str) -> str:
    return f"=MIN({data_range})"


def _column_data_range(column: int, last_data_row: int) -> str:
    col_letter = get_column_letter(column)
    return f"{col_letter}2:{col_letter}{last_data_row}"


def _rect_data_range(
    first_column: int,
    last_column: int,
    last_data_row: int,
) -> str:
    first_letter = get_column_letter(first_column)
    last_letter = get_column_letter(last_column)
    return f"{first_letter}2:{last_letter}{last_data_row}"


def _append_single_stats_table(
    worksheet: Worksheet,
    *,
    stats_start_row: int,
    start_col: int,
    data_range: str,
    title: str | None = None,
    usl: float | None = None,
    lsl: float | None = None,
) -> int:
    """Write one SPC stats block.

    When usl/lsl are given, col 2 holds recipe limits and col 3 USL/LSL
    are computed from the table's data range. U/T also use col 2 limits; CPK formulas use col 2.
    """
    label_col = start_col
    input_col = start_col + 1
    value_col = start_col + 2
    red_label = Font(color="FF0000")
    content_start_row = stats_start_row + (1 if title else 0)

    if title:
        worksheet.cell(row=stats_start_row, column=label_col, value=title)

    usl_row = content_start_row
    lsl_row = content_start_row + 1
    u_row = content_start_row + 2
    t_row = content_start_row + 3
    mean_row = content_start_row + 4
    stdev_row = content_start_row + 5
    max_row = content_start_row + 6
    min_row = content_start_row + 7
    cpku_row = content_start_row + 8
    cpkl_row = content_start_row + 9
    cpk_row = content_start_row + 10
    cp_row = content_start_row + 12
    ca_row = content_start_row + 13
    cpk2_row = content_start_row + 14

    usl_input_ref = _cell_ref(usl_row, input_col)
    lsl_input_ref = _cell_ref(lsl_row, input_col)
    usl_value_ref = _cell_ref(usl_row, value_col)
    lsl_value_ref = _cell_ref(lsl_row, value_col)
    u_ref = _cell_ref(u_row, value_col)
    t_ref = _cell_ref(t_row, value_col)
    mean_ref = _cell_ref(mean_row, value_col)
    stdev_ref = _cell_ref(stdev_row, value_col)
    cpku_ref = _cell_ref(cpku_row, value_col)
    cpkl_ref = _cell_ref(cpkl_row, value_col)
    cp_ref = _cell_ref(cp_row, value_col)
    ca_ref = _cell_ref(ca_row, value_col)

    if usl is not None and lsl is not None:
        usl_input: str | float | None = usl
        lsl_input: str | float | None = lsl
        usl_value_formula = _max_formula(data_range)
        lsl_value_formula = _min_formula(data_range)
        usl_spec_ref = usl_input_ref
        lsl_spec_ref = lsl_input_ref
    else:
        usl_input = "数据最大值"
        lsl_input = "数据最小值"
        usl_value_formula = _max_formula(data_range)
        lsl_value_formula = _min_formula(data_range)
        usl_spec_ref = usl_value_ref
        lsl_spec_ref = lsl_value_ref

    mean_formula = _mean_formula(data_range)
    stdev_formula = _stdev_formula(data_range)
    max_formula = _max_formula(data_range)
    min_formula = _min_formula(data_range)

    rows: list[tuple[str, str | float | None, str | None, bool]] = [
        ("公差上限 USL", usl_input, usl_value_formula, False),
        ("公差下限 LSL", lsl_input, lsl_value_formula, False),
        ("规格中心 U", "(USL + LSL) / 2", f"=({usl_spec_ref}+{lsl_spec_ref})/2", False),
        ("规格公差 T", "USL - LSL", f"={usl_spec_ref}-{lsl_spec_ref}", False),
        ("X 平均值", "na", mean_formula, False),
        ("标准差 σ", "STDEV", stdev_formula, False),
        ("最大值", "MAX", max_formula, False),
        ("最小值", "Min", min_formula, False),
        ("CPKU", "(USL - X) / 3σ", f"=({usl_spec_ref}-{mean_ref})/(3*{stdev_ref})", False),
        ("CPKL", "(X - LSL) / 3σ", f"=({mean_ref}-{lsl_spec_ref})/(3*{stdev_ref})", False),
        ("CPK", "Min(CPKU, CPKL)", f"=MIN({cpku_ref},{cpkl_ref})", True),
        ("", None, None, False),
        (
            "Cp 离散趋势精确度",
            "(USL - LSL) / 6σ",
            f"=({usl_spec_ref}-{lsl_spec_ref})/(6*{stdev_ref})",
            False,
        ),
        ("Ca 集中趋势精确度", "(X - U) / (T / 2)", f"=({mean_ref}-{u_ref})/({t_ref}/2)", False),
        ("Cpk", "Cp * (1 - |Ca|)", f"={cp_ref}*(1-ABS({ca_ref}))", True),
    ]

    stats_end_row = content_start_row + len(rows) - 1
    current_row = content_start_row
    for label, hint, formula, is_red in rows:
        label_cell = worksheet.cell(row=current_row, column=label_col, value=label or None)
        if is_red:
            label_cell.font = red_label
        if hint is not None:
            worksheet.cell(row=current_row, column=input_col, value=hint)
        if formula is not None:
            worksheet.cell(row=current_row, column=value_col, value=formula)
        current_row += 1

    black_side = Side(style="thin", color="000000")
    stats_border = Border(
        left=black_side,
        right=black_side,
        top=black_side,
        bottom=black_side,
    )
    for row in range(stats_start_row, stats_end_row + 1):
        for col in range(label_col, value_col + 1):
            worksheet.cell(row=row, column=col).border = stats_border

    if title:
        worksheet.merge_cells(
            start_row=stats_start_row,
            start_column=label_col,
            end_row=stats_start_row,
            end_column=value_col,
        )

    return stats_end_row


def _spc_table_title(
    recipe_name: str | None,
    aggregate_label: str,
    metric_label: str,
) -> str:
    prefix = f"{recipe_name}-" if recipe_name else ""
    return f"{prefix}{aggregate_label}{metric_label}"


def _scaled_recipe_limits(
    usl: float | None,
    lsl: float | None,
    multiplier: float,
) -> tuple[float | None, float | None]:
    if usl is None or lsl is None:
        return None, None
    return usl * multiplier, lsl * multiplier


def _append_spc_stats_tables(
    worksheet: Worksheet,
    *,
    last_data_row: int,
    tier_total_col: int,
    batch_total_col: int,
    sample_first_col: int,
    sample_last_col: int,
    metric_label: str,
    recipe_name: str | None = None,
    recipe_usl: float | None = None,
    recipe_lsl: float | None = None,
) -> None:
    """Append three side-by-side SPC tables: tier, batch, and raw sample values."""
    stats_start_row = last_data_row + 2
    first_table_col = tier_total_col + 2
    table_stride = 4  # 3 columns + 1 gap

    tier_range = _column_data_range(tier_total_col, last_data_row)
    batch_range = _column_data_range(batch_total_col, last_data_row)
    sample_range = _rect_data_range(sample_first_col, sample_last_col, last_data_row)

    table_specs: list[tuple[str, str, float | None, float | None]] = [
        ("单打", tier_range, *_scaled_recipe_limits(recipe_usl, recipe_lsl, 12)),
        ("单批", batch_range, *_scaled_recipe_limits(recipe_usl, recipe_lsl, 6)),
        ("单值", sample_range, recipe_usl, recipe_lsl),
    ]
    for offset, (aggregate_label, data_range, usl, lsl) in enumerate(table_specs):
        _append_single_stats_table(
            worksheet,
            stats_start_row=stats_start_row,
            start_col=first_table_col + offset * table_stride,
            data_range=data_range,
            title=_spc_table_title(recipe_name, aggregate_label, metric_label),
            usl=usl,
            lsl=lsl,
        )


def _draft_sheet_header(
    *,
    enable_water_cut: bool,
    enable_round_bread: bool,
) -> list[str]:
    header = [
        "批次号",
        "名称",
        "温度(°C)",
        "重量(g)",
    ]
    if enable_round_bread:
        header.append("直径(mm)")
    else:
        header.extend(["长(mm)", "宽(mm)"])
    header.append("高(mm)")
    if enable_water_cut:
        header.append("水切宽度(mm)")
    header.append("时间")
    return header


def _draft_sheet_row(
    record: MeasurementResponse,
    *,
    enable_water_cut: bool,
    enable_round_bread: bool,
) -> list[object]:
    row: list[object] = [
        record.batch_id,
        record.sample_name,
        record.temperature,
        record.weight,
    ]
    if enable_round_bread:
        row.append(record.length)
    else:
        row.extend([record.length, record.width])
    row.append(record.height)
    if enable_water_cut:
        row.append(
            record.water_cut_width
            if record.record_type == "product"
            else ""
        )
    row.append(format_recorded_at(record.recorded_at))
    return row


def _append_draft_sheet(
    workbook: Workbook,
    records: list[MeasurementResponse],
    *,
    enable_water_cut: bool,
    enable_round_bread: bool,
) -> None:
    """追加与数据汇总页表格一致的底稿 Sheet，置于 workbook 首位。"""
    worksheet = workbook.create_sheet(
        title=sanitize_sheet_title(DRAFT_SHEET_TITLE),
        index=0,
    )
    worksheet.append(
        _draft_sheet_header(
            enable_water_cut=enable_water_cut,
            enable_round_bread=enable_round_bread,
        )
    )
    for record in sorted(records, key=lambda item: item.recorded_at, reverse=True):
        worksheet.append(
            _draft_sheet_row(
                record,
                enable_water_cut=enable_water_cut,
                enable_round_bread=enable_round_bread,
            )
        )


def _append_metric_sheet(
    workbook: Workbook,
    *,
    sheet_title: str,
    metric_label: str,
    records: list[MeasurementResponse],
    value_getter: Callable[[MeasurementResponse], str],
    recipe: RecipeBase | None = None,
    record_type: RecordType | None = None,
    metric_field: str | None = None,
) -> None:
    grouped = _group_records_by_batch(records)
    max_samples = max((len(batch_records) for _, batch_records in grouped), default=0)

    worksheet: Worksheet = workbook.create_sheet(title=sheet_title)
    header = ["批次号", "开始时间"]
    header.extend(f"{metric_label}{index}" for index in range(1, max_samples + 1))
    header.append(f"单批{metric_label}")
    header.append(f"单打{metric_label}")
    worksheet.append(header)

    batch_rows: list[tuple[int | None, str, list[str], float]] = []
    for batch_id, batch_records in grouped:
        values = [value_getter(record) for record in batch_records]
        total = sum(_parse_numeric(value) for value in values)
        batch_rows.append(
            (
                batch_id or None,
                _batch_start_time(batch_records),
                values,
                total,
            )
        )

    tier_totals = _tier_totals_from_batch_totals([total for *_, total in batch_rows])

    for (batch_id, start_time, values, total), tier_total in zip(
        batch_rows,
        tier_totals,
        strict=True,
    ):
        row: list[object] = [
            batch_id,
            start_time,
            *[_metric_cell_value(value) for value in values],
        ]
        if len(values) < max_samples:
            row.extend([None] * (max_samples - len(values)))
        row.append(_format_sum(total))
        row.append(_format_sum(tier_total) if tier_total is not None else None)
        worksheet.append(row)

    last_data_row = worksheet.max_row
    batch_total_col = 2 + max_samples + 1
    tier_total_col = batch_total_col + 1
    sample_first_col = 3
    sample_last_col = 2 + max_samples
    recipe_usl: float | None = None
    recipe_lsl: float | None = None
    if recipe is not None and record_type is not None and metric_field is not None:
        recipe_lsl, recipe_usl = _limits_for_metric(recipe, record_type, metric_field)
    _append_spc_stats_tables(
        worksheet,
        last_data_row=last_data_row,
        tier_total_col=tier_total_col,
        batch_total_col=batch_total_col,
        sample_first_col=sample_first_col,
        sample_last_col=sample_last_col,
        metric_label=metric_label,
        recipe_name=recipe.name if recipe is not None else None,
        recipe_usl=recipe_usl,
        recipe_lsl=recipe_lsl,
    )


def build_measurements_xlsx(
    records: list[MeasurementResponse],
    *,
    enable_water_cut: bool = False,
    recipe: RecipeBase | None = None,
) -> bytes:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    records_by_type: dict[RecordType, list[MeasurementResponse]] = {
        "product": [],
        "bottom": [],
        "middle": [],
    }
    for record in records:
        records_by_type[record.record_type].append(record)

    enable_round_bread = recipe.enable_round_bread if recipe is not None else False

    if records:
        _append_draft_sheet(
            workbook,
            records,
            enable_water_cut=enable_water_cut,
            enable_round_bread=enable_round_bread,
        )

    for record_type in ("product", "bottom", "middle"):
        typed_records = sorted(records_by_type[record_type], key=_record_sort_key)
        if not typed_records:
            continue

        type_label = RECORD_TYPE_LABELS[record_type]
        for field, short_label, _metric_header in _export_metric_fields(
            enable_round_bread=enable_round_bread,
        ):
            if not _should_include_metric(
                record_type,
                field,
                enable_water_cut=enable_water_cut,
                enable_round_bread=enable_round_bread,
            ):
                continue

            sheet_title = sanitize_sheet_title(f"{type_label}-{short_label}")
            _append_metric_sheet(
                workbook,
                sheet_title=sheet_title,
                metric_label=short_label,
                records=typed_records,
                value_getter=lambda record, metric_field=field: _metric_value(record, metric_field),
                recipe=recipe,
                record_type=record_type,
                metric_field=field,
            )

    if not workbook.sheetnames:
        summary = workbook.create_sheet(title="录入数据")
        summary.append(["提示"])
        summary.append(["当前筛选条件下没有可导出的数据"])

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
